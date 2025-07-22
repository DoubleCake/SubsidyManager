# services/subsidy_service.py
from datetime import datetime
from typing import List, Dict, Optional, Any
from pathlib import Path
import json
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from models.subsidy_model import SubsidyDAO
from models.subsidy_rule_dao import SubsidyRuleDAO


class SubsidyService:
    """
    统一服务层：补贴类型 + 互斥规则 + 导入导出
    所有数据库操作走单例 DAO
    """

    def __init__(self, db_path: str = 'family_subsidies.db'):
        self.subsidy_dao = SubsidyDAO(db_path)   # 单例
        self.rule_dao    = SubsidyRuleDAO()

    # -------------------------------------------------
    # 补贴类型 CRUD
    # -------------------------------------------------
    def add_subsidy(self, name: str, amount: float, year: int,
                    land_type: str = "", description: str = "",
                    is_exclusive: bool = False, is_active: bool = True) -> int:
        data = {
            "name": name,
            "amount": amount,
            "year": year,
            "land_type": land_type,
            "description": description,
            "is_mutual_exclusive": int(is_exclusive),
            "is_activate": int(is_active)
        }
        return self.subsidy_dao.create_subsidy_type(data)

    def update_subsidy(self, subsidy_id: int, **kwargs) -> bool:
        return self.subsidy_dao.update_subsidy(subsidy_id, kwargs)

    def delete_subsidy(self, subsidy_id: int) -> bool:
        return self.subsidy_dao.delete_subsidy(subsidy_id)

    def get_subsidy(self, subsidy_id: int) -> Optional[Dict]:
        return self.subsidy_dao.get_subsidy_by_id(subsidy_id)

    def get_all_subsidies(self, active_only: bool = True) -> List[Dict]:
        """返回 {id, name} 供下拉框"""
        rows = self.subsidy_dao.get_all_subsidies(active_only)
        return [{"id": r["id"], "name": r["name"]} for r in rows]

    # -------------------------------------------------
    # 互斥规则 CRUD
    # -------------------------------------------------
    def add_rule(self, name: str, a_id: str, b_id: str,
                 relation: str, description: str = "") -> int:
        return self.rule_dao.add_rule(name, a_id, b_id, relation, description)

    def update_rule(self, rule_id: int, **kwargs) -> bool:
        return self.rule_dao.update_rule(rule_id, **kwargs)

    def delete_rule(self, rule_id: int) -> bool:
        return self.rule_dao.delete_rule(rule_id)

    def list_rules(self, **filters) -> List[Dict]:
        return self.rule_dao.search_rules(**filters)

    # -------------------------------------------------
    # 导出/导入
    # -------------------------------------------------
    def export_subsidies_to_csv(self, file_path: str, active_only: bool = True) -> None:
        subsidies = self.subsidy_dao.get_all_subsidies(active_only)
        if not subsidies:
            raise ValueError("无数据可导出")

        headers = ["id", "name", "amount", "year", "land_type",
                   "is_mutual_exclusive", "is_activate", "description"]
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for r in subsidies:
                writer.writerow({
                    "id": r["id"],
                    "name": r["name"],
                    "amount": r["amount"],
                    "year": r["year"],
                    "land_type": r.get("land_type", ""),
                    "is_mutual_exclusive": bool(r.get("is_mutual_exclusive")),
                    "is_activate": bool(r.get("is_activate")),
                    "description": r.get("description", "")
                })

    def import_subsidies_from_csv(self, file_path: str) -> int:
        count = 0
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    self.add_subsidy(
                        name=row["name"],
                        amount=float(row["amount"]),
                        year=int(row["year"]),
                        land_type=row.get("land_type", ""),
                        description=row.get("description", ""),
                        is_exclusive=row.get("is_mutual_exclusive", "").lower() == "是",
                        is_active=row.get("is_activate", "").lower() == "是"
                    )
                    count += 1
                except Exception as e:
                    print(f"跳过行: {e} {row}")
        return count

    def export_subsidies_to_excel(self, file_path: str, active_only: bool = True) -> None:
        subsidies = self.subsidy_dao.get_all_subsidies(active_only)
        if not subsidies:
            raise ValueError("无数据可导出")

        wb = Workbook()
        ws = wb.active
        ws.title = "补贴类型"
        headers = ["ID", "名称", "金额", "年份", "土地类型", "互斥", "激活", "描述"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, r in enumerate(subsidies, 2):
            ws.cell(row=row_idx, column=1, value=r["id"])
            ws.cell(row=row_idx, column=2, value=r["name"])
            ws.cell(row=row_idx, column=3, value=r["amount"])
            ws.cell(row=row_idx, column=4, value=r["year"])
            ws.cell(row=row_idx, column=5, value=r.get("land_type", ""))
            ws.cell(row=row_idx, column=6, value="是" if r.get("is_mutual_exclusive") else "否")
            ws.cell(row=row_idx, column=7, value="是" if r.get("is_activate") else "否")
            ws.cell(row=row_idx, column=8, value=r.get("description", ""))

        for col in ws.columns:
            max_len = max([len(str(cell.value or "")) for cell in col])
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        wb.save(file_path)